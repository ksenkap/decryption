import openpyxl

f = open('Nomera.txt')
numbers_hashes = []
numbers = []
for line in f:
    numbers_hashes.append(line.split(":")[0])
    numbers.append(line.rstrip().split(":")[1])
f.close()
russian_big = 'АБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'
russian_small = 'абвгдежзийклмнопрстуфхцчшщъыьэюя'
#russian_small = 'абвгдежзийклмнопрстуфхцчшщъыьэюяабвгдежзийклмнопрстуфхцчшщъыьэюя'
#russian_big = 'АБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯАБВГДЕЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'
english_small = 'abcdefghijklmnopqrstuvwxyz'
wb = openpyxl.load_workbook(filename='student_v.1.13.xlsx')
sheet = wb['A2']
sheet.cell(row=1, column=4).value = "Сдвиг"
for i in range(2, 1002):
    number_hash = sheet.cell(row=i, column=1).value
    ind = numbers_hashes.index(number_hash)
    sheet.cell(row=i, column=1).value = numbers[ind]

    adress = sheet.cell(row=i, column=3).value
    email = sheet.cell(row=i, column=2).value
    key = adress.split()[-1][0]
    distance = russian_small.index(key) - 10
    sheet.cell(row=i, column=4).value = distance
    adress_new = ""
    email_new = ""
    for j in range(len(adress)):
        if adress[j] in russian_small:
            adress_new += russian_small[(russian_small.index(adress[j])-distance)%32]
        elif adress[j] in russian_big:
            adress_new += russian_big[(russian_big.index(adress[j])-distance)%32]
        else:
            adress_new += adress[j]
    sheet.cell(row=i, column=3).value = adress_new
    for j in range(len(email)):
        if email[j] in english_small:
            if distance >= 0:
                email_new += english_small[(english_small.index(email[j])-distance)%26]
            else:
                distance += 32
                email_new += english_small[(english_small.index(email[j])-distance)%26]
        else:
            email_new += email[j]
    sheet.cell(row=i, column=2).value = email_new

wb.save('Answer.xlsx')
wb.close()
